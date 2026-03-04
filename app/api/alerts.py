from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import logging
from fastapi import APIRouter, Depends, HTTPException, Query, Response, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.ext.asyncio import AsyncSession
import asyncio
import math
from sqlalchemy import desc
from pydantic import BaseModel
import csv
import io

from app.db.session import get_db
from app.db.async_session import get_async_db
from app.models.alert import Alert, AlertResponse, AlertUpdate, AlertStatus
from app.services.alert_service import alert_service, register_sse_client, unregister_sse_client, publish_test_alert, connected_clients
# 导入JWT用户信息相关功能
from app.models.user import UserInfo
from app.core.auth import get_current_user, get_current_user_optional
from app.utils.auth_helper import get_user_id_from_request, get_user_name_from_request, get_user_context

logger = logging.getLogger(__name__)

# 导入 openpyxl 用于 Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel 导出功能将不可用")

router = APIRouter()

# 常量定义
ALERT_NOT_FOUND_MSG = "预警记录不存在"

@router.get("/stream", description="实时报警SSE流")
async def alert_stream(request: Request):
    """
    创建SSE连接，用于实时推送报警信息。
    使用StreamingResponse实现更稳定的SSE流。
    """
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    logger.info(f"收到SSE连接请求，客户端IP: {client_ip}")
    
    # 注册客户端
    client_queue = await register_sse_client(client_ip, user_agent)
    client_id = getattr(client_queue, '_client_id', 'unknown')
    logger.info(f"已注册SSE客户端，客户端ID: {client_id}")
    
    async def generate():
        try:
            # 发送连接成功消息
            yield "data: {\"event\": \"connected\"}\n\n"
            logger.info(f"SSE连接建立成功，客户端ID: {client_id}")
            
            while True:
                try:
                    # 检查客户端是否断开
                    if await request.is_disconnected():
                        logger.info(f"客户端断开连接，客户端ID: {client_id}")
                        break
                    
                    # 等待消息，超时则发送心跳
                    message = await asyncio.wait_for(client_queue.get(), timeout=10.0)
                    yield message
                    logger.debug(f"发送消息给客户端 {client_id}")
                    
                except asyncio.TimeoutError:
                    # 发送心跳
                    yield ": heartbeat\n\n"
                    
                except Exception as e:
                    logger.error(f"SSE流生成错误: {e}")
                    break
                    
        except Exception as e:
            logger.error(f"SSE连接异常: {e}")
        finally:
            # 清理客户端
            unregister_sse_client(client_queue)
            logger.info(f"SSE客户端已清理，客户端ID: {client_id}")
    
    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
        }
    )

@router.get("/real-time")  # 向后兼容的路由
async def get_realtime_alerts(
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1, description="页码"),
    limit: int = Query(10, ge=1, le=100, description="每页数量"),
    alert_type: Optional[str] = Query(None, description="报警类型"),
    camera_id: Optional[int] = Query(None, description="摄像头ID"),
    camera_name: Optional[str] = Query(None, description="摄像头名称"),
    alert_level: Optional[int] = Query(None, description="报警等级"),
    alert_name: Optional[str] = Query(None, description="报警名称"),
    task_id: Optional[int] = Query(None, description="任务ID"),
    location: Optional[str] = Query(None, description="位置"),
    status: Optional[str] = Query(None, description="报警状态：1=待处理，2=处理中，3=已处理，4=已忽略，5=已过期"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    start_time: Optional[str] = Query(None, description="开始时间 (ISO格式)"),
    end_time: Optional[str] = Query(None, description="结束时间 (ISO格式)"),
    skill_class_id: Optional[int] = Query(None, description="技能类别ID"),
    alert_id: Optional[int] = Query(None, description="报警ID")
):
    """
    获取实时预警列表，支持分页和多维度过滤
    
    🎯 企业级筛选功能：
    - 状态筛选：支持按报警处理状态筛选
    - 日期范围筛选：支持按预警时间的开始日期和结束日期筛选  
    - 多维度过滤：摄像头、类型、等级、位置等
    - 高性能分页：支持大数据量场景
    """
    logger.info(f"收到获取实时预警列表请求: camera_id={camera_id}, camera_name={camera_name}, " 
               f"alert_type={alert_type}, alert_level={alert_level}, alert_name={alert_name}, "
               f"task_id={task_id}, location={location}, status={status}, "
               f"start_date={start_date}, end_date={end_date}, start_time={start_time}, end_time={end_time}, "
               f"skill_class_id={skill_class_id}, alert_id={alert_id}, "
               f"page={page}, limit={limit}")
    
    # 🚀 参数验证和转换
    try:
        # 转换日期字符串为datetime对象
        parsed_start_date = None
        parsed_end_date = None
        parsed_start_time = None  
        parsed_end_time = None
        
        if start_date:
            try:
                parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d")
                logger.debug(f"解析开始日期: {start_date} -> {parsed_start_date}")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"开始日期格式错误，应为YYYY-MM-DD格式: {start_date}")
        
        if end_date:
            try:
                # 结束日期设置为当天的23:59:59
                parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                logger.debug(f"解析结束日期: {end_date} -> {parsed_end_date}")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"结束日期格式错误，应为YYYY-MM-DD格式: {end_date}")
        
        if start_time:
            try:
                parsed_start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                logger.debug(f"解析开始时间: {start_time} -> {parsed_start_time}")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"开始时间格式错误，应为ISO格式: {start_time}")
        
        if end_time:
            try:
                parsed_end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
                logger.debug(f"解析结束时间: {end_time} -> {parsed_end_time}")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"结束时间格式错误，应为ISO格式: {end_time}")
        
        # 验证状态值并转换为数字
        status_value = None
        if status:
            status_map = {
                "待处理": 1,
                "处理中": 2,
                "已处理": 3,
                "已忽略": 4,
                "已过期": 5
            }
            if status not in status_map:
                raise HTTPException(status_code=400, detail=f"无效的状态值: {status}")
            status_value = status_map[status]
            
        # 验证日期范围
        if parsed_start_date and parsed_end_date and parsed_start_date > parsed_end_date:
            raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
            
        if parsed_start_time and parsed_end_time and parsed_start_time > parsed_end_time:
            raise HTTPException(status_code=400, detail="开始时间不能晚于结束时间")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"参数解析失败: {str(e)}")
        raise HTTPException(status_code=400, detail=f"参数解析失败: {str(e)}")
    
    # 计算分页跳过的记录数
    skip = (page - 1) * limit
    
    # 🆕 应用筛选条件
    filtered_alerts = await alert_service.get_alerts(
        db=db,
        skip=skip,
        limit=limit,
        alert_type=alert_type,
        camera_id=camera_id,
        camera_name=camera_name,
        alert_level=alert_level,
        alert_name=alert_name,
        task_id=task_id,
        location=location,
        status=status_value,
        start_date=start_date,
        end_date=end_date,
        start_time=start_time,
        end_time=end_time,
        skill_class_id=skill_class_id,
        alert_id=alert_id
    )
    
    # 🆕 获取总数（应用相同的筛选条件）
    total_count = await alert_service.get_alerts_count(
        db=db,
        alert_type=alert_type,
        camera_id=camera_id,
        camera_name=camera_name,
        alert_level=alert_level,
        alert_name=alert_name,
        task_id=task_id,
        location=location,
        status=status_value,
        start_date=start_date,
        end_date=end_date,
        start_time=start_time,
        end_time=end_time,
        skill_class_id=skill_class_id,
        alert_id=alert_id
    )
    
    # 计算总页数
    try:
        pages = math.ceil(total_count / limit)
    except (TypeError, ValueError):
        # 处理无法转换为整数的情况
        pages = 1
    
    # 将Alert对象转换为AlertResponse对象
    alert_responses = [AlertResponse.model_validate(alert) for alert in filtered_alerts]

    logger.info(f"获取实时预警列表成功，返回 {len(alert_responses)} 条记录，总共 {total_count} 条")

    # 使用统一响应格式
    return {
        "success": True,
        "code": 200,
        "message": "获取预警列表成功",
        "data": alert_responses,
        "pagination": {
            "total": total_count,
            "page": page,
            "page_size": limit,
            "pages": pages,
            "has_next": page < pages,
            "has_prev": page > 1
        }
    }

@router.get("/sse/status", description="获取SSE连接状态")
def get_sse_status():
    """
    获取SSE连接状态信息
    """
    try:
        logger.info("收到获取SSE状态请求")
        
        # 获取连接管理器状态
        from app.services.sse_connection_manager import sse_manager
        
        status_info = {
            "success": True,
            "sse_enabled": True,
            "total_connections": len(connected_clients),
            "manager_status": {
                "is_running": sse_manager.is_running if hasattr(sse_manager, 'is_running') else True,
                "start_time": getattr(sse_manager, 'start_time', None),
                "total_messages_sent": getattr(sse_manager, 'total_messages_sent', 0),
                "active_connections": getattr(sse_manager, 'active_connections', len(connected_clients))
            },
            "performance": {
                "queue_size_limit": getattr(sse_manager, 'queue_size_limit', 1000),
                "send_timeout": getattr(sse_manager, 'send_timeout', 2.0),
                "batch_size": getattr(sse_manager, 'batch_size', 10)
            }
        }
        
        return status_info
    except Exception as e:
        logger.error(f"获取SSE状态失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取SSE状态失败: {str(e)}")

@router.get("/statistics", description="获取报警统计信息")
async def get_alert_statistics(
    db: AsyncSession = Depends(get_async_db),
    days: int = Query(30, ge=1, le=3650, description="统计天数，设为3650可统计约10年数据"),
    all_time: bool = Query(False, description="是否统计全部历史数据（忽略时间范围）")
):
    """
    获取报警统计信息

    参数:
    - days: 统计最近N天的数据（默认30天，最大3650天约10年）
    - all_time: 为true时统计全部历史数据，忽略days参数

    返回:
    - 预警总数
    - 各状态数量
    - 预警类型排名
    - 预警等级分布
    - 按天统计
    - 设备预警数量 Top 10
    """
    try:
        logger.info(f"收到获取报警统计请求，统计天数: {days}, 全部历史: {all_time}")

        # 计算时间范围
        from datetime import datetime, timedelta
        end_date = datetime.now()

        if all_time:
            # 统计全部历史数据，使用一个极早的起始时间
            start_date = datetime(2000, 1, 1)
        else:
            start_date = end_date - timedelta(days=days)

        # 获取统计数据
        stats = await alert_service.get_alert_statistics(
            db=db,
            start_date=start_date,
            end_date=end_date
        )

        return {
            "success": True,
            "time_range": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "days": days
            },
            "statistics": stats
        }
    except Exception as e:
        logger.error(f"获取报警统计失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取报警统计失败: {str(e)}")

@router.get("/latest-images", description="获取最新预警图片（用于大屏展示）")
async def get_latest_alert_images(
    limit: int = Query(10, ge=1, le=50, description="返回数量限制"),
    db: Session = Depends(get_db)
):
    """
    获取最新预警图片列表

    返回最新的预警记录，包含图片URL、预警信息等，
    主要用于大屏可视化中心的预警图片轮播展示。

    级别映射：
    - urgent: 一级（紧急）
    - high: 二级（重要）
    - medium: 三级（普通）
    - low: 四级（提示）
    """
    try:
        logger.info(f"收到获取最新预警图片请求，限制数量: {limit}")

        # 级别映射
        level_map = {
            4: ("urgent", "一级"),
            3: ("high", "二级"),
            2: ("medium", "三级"),
            1: ("low", "四级")
        }

        # 查询最新有图片的预警记录
        query = db.query(Alert).filter(
            Alert.minio_frame_object_name.isnot(None),
            Alert.minio_frame_object_name != ""
        ).order_by(desc(Alert.alert_time)).limit(limit)

        alerts = query.all()

        # 构建返回数据
        result = []
        for alert in alerts:
            # 获取级别信息
            level_code, level_text = level_map.get(alert.alert_level, (1, "四级"))

            # 使用AlertResponse模型自动生成minio_frame_url
            alert_response = AlertResponse.model_validate(alert)

            result.append({
                "id": alert.alert_id,
                "image_url": alert_response.minio_frame_url,
                "alert_type": alert.alert_name,
                "alert_time": alert.alert_time.strftime("%Y-%m-%d %H:%M") if alert.alert_time else "",
                "level": level_code,
                "levelText": level_text,
                "location": alert.location or "",
                "camera_name": alert.camera_name or ""
            })

        logger.info(f"成功获取最新预警图片 {len(result)} 条")

        return {
            "code": 0,
            "msg": "success",
            "data": result
        }

    except Exception as e:
        logger.error(f"获取最新预警图片失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取最新预警图片失败: {str(e)}")



# ========== 预警统计API端点 ==========

@router.get("/statistics/summary", summary="获取预警统计摘要")
async def get_summary_statistics(
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """
    获取预警统计摘要
    
    返回信息包括：
    - total_alerts: 总预警数
    - today_alerts: 今日新增
    - pending_alerts: 待处理数
    - processing_alerts: 处理中数
    - resolved_today: 今日已处理
    - total_cameras: 总设备数
    - online_cameras: 在线设备数
    - offline_cameras: 离线设备数
    """
    try:
        logger.info(f"收到获取预警统计摘要请求: start_date={start_date}, end_date={end_date}")
        
        # 解析日期参数
        parsed_start_date = None
        parsed_end_date = None
        
        if start_date:
            try:
                parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"开始日期格式错误: {start_date}")
        
        if end_date:
            try:
                parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d")
                parsed_end_date = parsed_end_date.replace(hour=23, minute=59, second=59)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"结束日期格式错误: {end_date}")
        
        # 获取统计数据
        stats = alert_service.get_summary_stats(
            db=db,
            start_date=parsed_start_date,
            end_date=parsed_end_date
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警统计摘要失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警统计摘要失败: {str(e)}")


@router.get("/statistics/trend", summary="获取预警趋势数据")
async def get_trend_statistics(
    time_range: str = Query("7d", description="时间范围: 24h/7d/30d"),
    granularity: str = Query("day", description="时间粒度: hour/day/week/month"),
    db: Session = Depends(get_db)
):
    """
    获取预警趋势数据

    支持的时间范围：
    - 24h / day: 最近24小时
    - 7d / week: 最近7天
    - 30d / month: 最近30天

    支持的时间粒度：
    - hour: 按小时统计
    - day: 按天统计
    - week: 按周统计
    - month: 按月统计
    """
    try:
        logger.info(f"收到获取预警趋势请求: time_range={time_range}, granularity={granularity}")

        # 解析时间范围
        end_date = datetime.now()
        range_mapping = {
            "24h": timedelta(hours=24),
            "day": timedelta(hours=24),
            "7d": timedelta(days=7),
            "week": timedelta(days=7),
            "30d": timedelta(days=30),
            "month": timedelta(days=30)
        }

        delta = range_mapping.get(time_range, timedelta(days=7))
        start_date = end_date - delta
        
        # 获取趋势数据
        stats = alert_service.get_trend_stats(
            db=db,
            start_date=start_date,
            end_date=end_date,
            granularity=granularity
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警趋势失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警趋势失败: {str(e)}")


@router.get("/statistics/by-type", summary="获取预警类型统计")
async def get_type_statistics(
    range: str = Query("7d", description="时间范围: 24h/7d/30d"),
    top_n: int = Query(10, ge=1, le=50, description="返回前N个类型"),
    db: Session = Depends(get_db)
):
    """
    获取预警类型统计

    返回各预警类型的数量和占比
    """
    try:
        logger.info(f"收到获取预警类型统计请求: range={range}, top_n={top_n}")

        # 解析时间范围
        end_date = datetime.now()
        range_mapping = {
            "24h": timedelta(hours=24),
            "day": timedelta(hours=24),
            "7d": timedelta(days=7),
            "week": timedelta(days=7),
            "30d": timedelta(days=30),
            "month": timedelta(days=30)
        }

        delta = range_mapping.get(range, timedelta(days=7))
        start_date = end_date - delta
        
        # 获取类型统计
        stats = alert_service.get_type_stats(
            db=db,
            start_date=start_date,
            end_date=end_date,
            top_n=top_n
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警类型统计失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警类型统计失败: {str(e)}")


@router.get("/statistics/level", summary="获取预警等级统计")
async def get_level_statistics(
    range: str = Query("7d", description="时间范围: 24h/7d/30d"),
    db: Session = Depends(get_db)
):
    """
    获取预警等级统计

    返回各预警等级的数量、占比和颜色
    """
    try:
        logger.info(f"收到获取预警等级统计请求: range={range}")

        # 解析时间范围
        end_date = datetime.now()
        range_mapping = {
            "24h": timedelta(hours=24),
            "day": timedelta(hours=24),
            "7d": timedelta(days=7),
            "week": timedelta(days=7),
            "30d": timedelta(days=30),
            "month": timedelta(days=30)
        }

        delta = range_mapping.get(range, timedelta(days=7))
        start_date = end_date - delta
        
        # 获取等级统计
        stats = alert_service.get_level_stats(
            db=db,
            start_date=start_date,
            end_date=end_date
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警等级统计失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警等级统计失败: {str(e)}")


@router.get("/statistics/location", summary="获取位置统计")
async def get_location_statistics(
    range: str = Query("7d", description="时间范围: 24h/7d/30d"),
    top_n: int = Query(10, ge=1, le=50, description="返回前N个位置"),
    db: Session = Depends(get_db)
):
    """
    获取按位置（摄像头）统计的预警数据

    返回各摄像头的预警数量和占比
    """
    try:
        logger.info(f"收到获取位置统计请求: range={range}, top_n={top_n}")

        # 解析时间范围
        end_date = datetime.now()
        range_mapping = {
            "24h": timedelta(hours=24),
            "day": timedelta(hours=24),
            "7d": timedelta(days=7),
            "week": timedelta(days=7),
            "30d": timedelta(days=30),
            "month": timedelta(days=30)
        }

        delta = range_mapping.get(range, timedelta(days=7))
        start_date = end_date - delta
        
        # 获取位置统计
        stats = alert_service.get_location_stats(
            db=db,
            start_date=start_date,
            end_date=end_date,
            top_n=top_n
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取位置统计失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取位置统计失败: {str(e)}")


@router.get("/statistics/processing-status", summary="获取处理状态统计")
async def get_processing_status_statistics(
    range: str = Query("7d", description="时间范围: 24h/7d/30d"),
    db: Session = Depends(get_db)
):
    """
    获取处理状态统计

    返回各处理状态的预警数量和占比
    """
    try:
        logger.info(f"收到获取处理状态统计请求: range={range}")

        # 解析时间范围
        end_date = datetime.now()
        range_mapping = {
            "24h": timedelta(hours=24),
            "day": timedelta(hours=24),
            "7d": timedelta(days=7),
            "week": timedelta(days=7),
            "30d": timedelta(days=30),
            "month": timedelta(days=30)
        }

        delta = range_mapping.get(range, timedelta(days=7))
        start_date = end_date - delta
        
        # 获取处理状态统计
        stats = alert_service.get_processing_status_stats(
            db=db,
            start_date=start_date,
            end_date=end_date
        )
        
        return {
            "code": 0,
            "msg": "success",
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取处理状态统计失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取处理状态统计失败: {str(e)}")

@router.get("/connected")
def get_connected_clients():
    """
    获取当前连接的SSE客户端信息
    """
    try:
        logger.info("收到获取连接客户端信息请求")
        clients_info = []
        
        # connected_clients 是一个set，包含客户端队列对象
        for client_queue in connected_clients:
            client_info = {
                "client_id": getattr(client_queue, '_client_id', f"client_{id(client_queue)}"),
                "connection_time": getattr(client_queue, '_connection_time', None),
                "queue_size": client_queue.qsize() if hasattr(client_queue, 'qsize') else 0,
                "client_ip": getattr(client_queue, '_client_ip', 'unknown'),
                "user_agent": getattr(client_queue, '_user_agent', 'unknown'),
                "is_connected": True  # 如果在set中说明连接是活跃的
            }
            clients_info.append(client_info)
        
        return {
            "success": True,
            "total_clients": len(connected_clients),
            "clients": clients_info
        }
    except Exception as e:
        logger.error(f"获取连接客户端信息失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取连接客户端信息失败: {str(e)}")


@router.get("/export", summary="导出预警数据")
async def export_alerts(
    db: Session = Depends(get_db),
    format: str = Query("csv", description="导出格式: csv 或 excel"),
    alert_ids: Optional[List[int]] = Query(None, description="指定导出的预警ID列表"),
    alert_type: Optional[str] = Query(None, description="报警类型"),
    camera_id: Optional[int] = Query(None, description="摄像头ID"),
    camera_name: Optional[str] = Query(None, description="摄像头名称"),
    alert_level: Optional[int] = Query(None, description="报警等级"),
    alert_name: Optional[str] = Query(None, description="报警名称"),
    task_id: Optional[int] = Query(None, description="任务ID"),
    location: Optional[str] = Query(None, description="位置"),
    status: Optional[str] = Query(None, description="报警状态"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    start_time: Optional[str] = Query(None, description="开始时间 (ISO格式)"),
    end_time: Optional[str] = Query(None, description="结束时间 (ISO格式)"),
    skill_class_id: Optional[int] = Query(None, description="技能类别ID"),
    alert_id: Optional[int] = Query(None, description="报警ID")
):
    """
    导出预警数据
    
    支持根据筛选条件导出CSV或Excel格式的预警数据
    如果指定了alert_ids，则只导出这些预警；否则根据筛选条件导出所有匹配的预警
    """
    try:
        logger.info(f"收到导出预警数据请求: format={format}, alert_ids={alert_ids}, "
                   f"camera_id={camera_id}, camera_name={camera_name}, alert_type={alert_type}")
        
        # 验证导出格式
        if format not in ['csv', 'excel']:
            raise HTTPException(status_code=400, detail=f"不支持的导出格式: {format}，仅支持 csv 或 excel")
        
        # 验证日期参数格式
        if start_date:
            try:
                datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"开始日期格式错误: {start_date}")
        
        if end_date:
            try:
                datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"结束日期格式错误: {end_date}")
        
        if start_time:
            try:
                datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"开始时间格式错误: {start_time}")
        
        if end_time:
            try:
                datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"结束时间格式错误: {end_time}")
        
        # 转换状态值
        status_value = None
        if status:
            status_map = {
                "待处理": 1,
                "处理中": 2,
                "已处理": 3,
                "已忽略": 4,
                "已过期": 5
            }
            if status in status_map:
                status_value = status_map[status]
        
        # 如果指定了alert_ids，直接查询这些预警
        if alert_ids:
            logger.info(f"导出指定的 {len(alert_ids)} 个预警")
            alerts = db.query(Alert).filter(Alert.alert_id.in_(alert_ids)).order_by(desc(Alert.alert_time)).all()
        else:
            # 否则根据筛选条件查询所有匹配的预警（不限制数量）
            logger.info("导出所有筛选条件下的预警（无数量限制）")
            alerts = await alert_service.get_alerts(
                db=db,
                skip=0,
                limit=999999,  # 设置一个非常大的限制，实际上等于不限制
                alert_type=alert_type,
                camera_id=camera_id,
                camera_name=camera_name,
                alert_level=alert_level,
                alert_name=alert_name,
                task_id=task_id,
                location=location,
                status=status_value,
                start_date=start_date,
                end_date=end_date,
                start_time=start_time,
                end_time=end_time,
                skill_class_id=skill_class_id,
                alert_id=alert_id
            )
        
        if not alerts:
            raise HTTPException(status_code=404, detail="没有找到符合条件的预警数据")
        
        logger.info(f"找到 {len(alerts)} 条预警数据，开始生成{format.upper()}文件")
        
        # 生成CSV文件
        if format == 'csv':
            # 创建CSV内容
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 写入表头
            headers = [
                "预警ID", "预警名称", "预警类型", "预警等级", "摄像头名称", 
                "位置", "预警时间", "状态", "处理人", "处理时间", "处理备注"
            ]
            writer.writerow(headers)
            
            # 写入数据行
            for alert in alerts:
                # 状态显示名称
                status_display = AlertStatus.get_display_name(alert.status)
                
                # 预警等级显示
                level_display = f"等级{alert.alert_level}" if alert.alert_level else "-"
                
                # 格式化时间
                alert_time = alert.alert_time.strftime("%Y-%m-%d %H:%M:%S") if alert.alert_time else "-"
                processed_at = alert.processed_at.strftime("%Y-%m-%d %H:%M:%S") if alert.processed_at else "-"
                
                row = [
                    alert.alert_id,
                    alert.alert_name or "-",
                    alert.alert_type or "-",
                    level_display,
                    alert.camera_name or "-",
                    alert.location or "-",
                    alert_time,
                    status_display,
                    alert.processed_by or "-",
                    processed_at,
                    alert.processing_notes or "-"
                ]
                writer.writerow(row)
            
            # 获取CSV内容
            csv_content = output.getvalue()
            output.close()
            
            # 返回CSV文件
            logger.info(f"CSV文件生成完成，共 {len(alerts)} 条记录")
            return Response(
                content=csv_content.encode('utf-8-sig'),  # 使用utf-8-sig编码支持Excel打开中文
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename=alerts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                }
            )
        
        elif format == 'excel':
            # 检查 openpyxl 是否可用
            if not OPENPYXL_AVAILABLE:
                raise HTTPException(
                    status_code=501, 
                    detail="Excel格式导出功能需要安装 openpyxl 库，请运行: pip install openpyxl"
                )
            
            # 创建 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "预警数据"
            
            # 定义样式
            # 表头样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 数据样式
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            data_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            
            # 写入表头
            headers = [
                "预警ID", "预警名称", "预警类型", "预警等级", "摄像头名称", 
                "位置", "预警时间", "状态", "处理人", "处理时间", "处理备注"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # 写入数据行
            for row_num, alert in enumerate(alerts, 2):
                # 状态显示名称
                status_display = AlertStatus.get_display_name(alert.status)
                
                # 预警等级显示
                level_display = f"等级{alert.alert_level}" if alert.alert_level else "-"
                
                # 格式化时间
                alert_time = alert.alert_time.strftime("%Y-%m-%d %H:%M:%S") if alert.alert_time else "-"
                processed_at = alert.processed_at.strftime("%Y-%m-%d %H:%M:%S") if alert.processed_at else "-"
                
                row_data = [
                    alert.alert_id,
                    alert.alert_name or "-",
                    alert.alert_type or "-",
                    level_display,
                    alert.camera_name or "-",
                    alert.location or "-",
                    alert_time,
                    status_display,
                    alert.processed_by or "-",
                    processed_at,
                    alert.processing_notes or "-"
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = data_alignment
                    cell.border = data_border
            
            # 调整列宽
            column_widths = {
                'A': 10,  # 预警ID
                'B': 25,  # 预警名称
                'C': 20,  # 预警类型
                'D': 12,  # 预警等级
                'E': 20,  # 摄像头名称
                'F': 20,  # 位置
                'G': 20,  # 预警时间
                'H': 12,  # 状态
                'I': 15,  # 处理人
                'J': 20,  # 处理时间
                'K': 30,  # 处理备注
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 冻结首行（表头）
            ws.freeze_panes = 'A2'
            
            # 保存到内存
            excel_output = io.BytesIO()
            wb.save(excel_output)
            excel_output.seek(0)
            
            # 返回 Excel 文件
            logger.info(f"Excel文件生成完成，共 {len(alerts)} 条记录")
            return Response(
                content=excel_output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=alerts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                }
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出预警数据失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出预警数据失败: {str(e)}")


@router.get("/{alert_id}", response_model=AlertResponse)
def get_alert(
    alert_id: int,
    db: Session = Depends(get_db)
):
    """
    根据ID获取单个报警记录详情，包含完整的处理流程信息
    """
    logger.info(f"收到获取报警详情请求: ID={alert_id}")
    
    alert = alert_service.get_alert_by_id(db, str(alert_id))
    if alert is None:
        logger.warning(f"报警记录不存在: ID={alert_id}")
        raise HTTPException(status_code=404, detail="报警记录不存在")
    
    # 🆕 使用AlertResponse.model_validate转换，确保包含所有字段和URL
    alert_response = AlertResponse.model_validate(alert)
    
    logger.info(f"获取报警详情成功: ID={alert_id}, 处理步骤数: {len(alert_response.process.get('steps', [])) if alert_response.process else 0}")
    return alert_response



@router.get("/{alert_id}/process", response_model=Dict[str, Any])
def get_alert_process(
    alert_id: int,
    db: Session = Depends(get_db)
):
    """
    获取报警的处理流程详情
    """
    logger.info(f"收到获取报警处理流程请求: ID={alert_id}")
    
    alert = alert_service.get_alert_by_id(db, str(alert_id))
    if alert is None:
        logger.warning(f"报警记录不存在: ID={alert_id}")
        raise HTTPException(status_code=404, detail="报警记录不存在")
    
    # 获取处理流程信息
    process_info = alert.process or {"remark": "", "steps": []}
    process_summary = alert.get_process_summary()
    
    response = {
        "alert_id": alert.alert_id,
        "current_status": alert.status,
        "current_status_display": AlertStatus.get_display_name(alert.status),
        "process": process_info,
        "summary": process_summary
    }
    
    logger.info(f"获取报警处理流程成功: ID={alert_id}, 步骤数: {process_summary['total_steps']}")
    return response

@router.post("/test", description="发送测试报警（高性能优化版本）")
async def send_test_alert(
    db: Session = Depends(get_db)
):
    """
    🚀 高性能测试报警接口 - 异步处理优化
    
    优化策略：
    1. 快速响应：接口立即返回，后台异步处理
    2. 异步MinIO上传：避免IO阻塞
    3. 数据库查询缓存：减少重复查询
    """
    logger.info("收到发送测试报警请求 - 高性能版本")
    
    try:
        # 导入必要的模块
        from app.services.ai_task_executor import task_executor
        from app.models.ai_task import AITask
        import numpy as np
        import cv2
        import json
        import asyncio
        from datetime import datetime
        
        # 🚀 优化1：预构建轻量级模拟数据
        mock_task = AITask(
            id=9999, name="测试报警任务", description="高性能测试", status=True,
            alert_level=1, frame_rate=1.0, task_type="detection", config='{}',
            camera_id=123, skill_class_id=9999, skill_config='{}',
            running_period='{"enabled": true, "periods": [{"start": "00:00", "end": "23:59"}]}',
            electronic_fence='{"enabled": true, "points": [[{"x": 100, "y": 80}, {"x": 500, "y": 80}, {"x": 500, "y": 350}, {"x": 100, "y": 350}]], "trigger_mode": "inside"}'
        )
        
        # 🚀 优化2：简化报警数据结构
        mock_alert_data = {
            "detections": [
                {"bbox": [383, 113, 472, 317], "confidence": 0.82, "class_name": "果蔬生鲜"},
                {"bbox": [139, 105, 251, 308], "confidence": 0.86, "class_name": "家居家纺"},
                {"bbox": [491, 125, 558, 301], "confidence": 0.62, "class_name": "食品饮料"}
            ],
            "alert_info": {
                "alert_triggered": True, "alert_level": 1,
                "alert_name": "商品区域检测报警", "alert_type": "product_area_detection",
                "alert_description": "检测到多个商品区域有异常活动，请及时查看"
            }
        }
        
        # 创建模拟的图像帧（640x480的蓝色图像，标准监控摄像头分辨率）
        mock_frame = np.full((480, 640, 3), (255, 128, 0), dtype=np.uint8)  # 橙蓝色背景
        
        # 绘制多个检测框和标签
        # 1. 果蔬生鲜区域（绿色框）
        cv2.rectangle(mock_frame, (383, 113), (472, 317), (0, 255, 0), 2)
        cv2.putText(mock_frame, "果蔬生鲜 0.82", (385, 110), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        
        # 2. 家居家纺区域（蓝色框）
        cv2.rectangle(mock_frame, (139, 105), (251, 308), (255, 0, 0), 2)
        cv2.putText(mock_frame, "家居家纺 0.86", (141, 102), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
        
        # 3. 食品饮料区域（红色框）
        cv2.rectangle(mock_frame, (491, 125), (558, 301), (0, 0, 255), 2)
        cv2.putText(mock_frame, "食品饮料 0.62", (493, 122), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        
        # 在左上角添加时间戳和摄像头信息
        timestamp_text = f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cv2.putText(mock_frame, timestamp_text, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.putText(mock_frame, "摄像头ID: 123", (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        
        # 🚀 优化6：异步处理 - 立即返回响应，后台处理
        task_id = f"test_{int(datetime.now().timestamp())}"
        
        # 创建异步任务，不等待完成
        async def process_alert_async():
            try:
                result = await asyncio.get_event_loop().run_in_executor(
                    task_executor.alert_executor,  # 使用现有线程池
                    task_executor._generate_alert_async_optimized,  # 新的优化方法
                    mock_task, mock_alert_data, mock_frame, 1
                )
                if result:
                    logger.info(f"✅ 异步测试报警处理完成: task_id={task_id}")
                else:
                    logger.warning(f"⚠️ 异步测试报警处理失败: task_id={task_id}")
            except Exception as e:
                logger.error(f"❌ 异步测试报警处理异常: task_id={task_id}, error={e}")
        
        # 启动异步任务（fire-and-forget）
        asyncio.create_task(process_alert_async())
        
        # 🚀 立即返回响应（不等待MinIO上传）
        logger.info(f"✅ 测试报警请求已接收并进入异步处理队列: task_id={task_id}")
        return {
            "success": True,
            "message": "测试报警已进入处理队列，正在后台异步处理",
            "task_id": task_id,
            "method": "async_optimized",
            "optimization": {
                "async_processing": True,
                "database_cache": "摄像头和技能信息缓存5分钟",
                "fast_response": "立即返回，后台处理",
                "expected_improvement": "响应时间从数秒降至数十毫秒"
            }
        }
            
    except Exception as e:
        logger.error(f"发送测试报警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"发送测试报警失败: {str(e)}")


# ========== 预警处理增强功能 ==========

@router.post("/{alert_id}/start-processing", response_model=AlertResponse, description="开始处理预警（确认处理）")
def start_processing_alert(
    alert_id: int,
    processing_notes: str,
    processed_by: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    开始处理预警（确认处理功能）
    将预警状态更新为"处理中"并记录处理意见
    """
    try:
        # 构建状态更新请求
        status_update = AlertUpdate(
            status=AlertStatus.PROCESSING,
            processed_by=processed_by,
            processing_notes=processing_notes
        )
        
        updated_alert = alert_service.update_alert_status(db, alert_id, status_update)
        if not updated_alert:
            raise HTTPException(status_code=404, detail="预警记录不存在")
        
        alert_response = AlertResponse.model_validate(updated_alert)
        logger.info(f"✅ 成功开始处理预警 {alert_id}，处理人: {processed_by}")
        return alert_response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"开始处理预警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"开始处理预警失败: {str(e)}")


@router.post("/{alert_id}/finish-processing", response_model=AlertResponse, description="完成处理预警（结束处理）")
def finish_processing_alert(
    alert_id: int,
    final_notes: Optional[str] = None,
    processed_by: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    完成处理预警（结束处理功能）
    将预警状态更新为"已处理"并记录完成信息
    """
    try:
        # 检查当前状态
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail="预警记录不存在")
        
        if alert.status != AlertStatus.PROCESSING:
            raise HTTPException(
                status_code=400, 
                detail=f"预警当前状态为{AlertStatus.get_display_name(alert.status)}，不能完成处理"
            )

        # 构建状态更新请求
        final_processing_notes = final_notes or "处理已完成"
        status_update = AlertUpdate(
            status=AlertStatus.RESOLVED,
            processed_by=processed_by,
            processing_notes=final_processing_notes
        )
        
        updated_alert = alert_service.update_alert_status(db, alert_id, status_update)
        alert_response = AlertResponse.model_validate(updated_alert)
        
        logger.info(f"✅ 成功完成处理预警 {alert_id}，处理人: {processed_by}")
        return alert_response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"完成处理预警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"完成处理预警失败: {str(e)}")


@router.get("/{alert_id}/processing-history", description="获取预警处理历史")
def get_alert_processing_history(
    alert_id: int,
    db: Session = Depends(get_db)
):
    """
    获取预警处理历史，解析process字段中的步骤信息
    """
    try:
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail="预警记录不存在")
        
        processing_history = []
        if alert.process and 'steps' in alert.process:
            for step in alert.process['steps']:
                processing_history.append({
                    "step": step.get('step', ''),
                    "time": step.get('time', ''),
                    "description": step.get('desc', ''),
                    "operator": step.get('operator', '')
                })
        
        result = {
            "alert_id": alert.alert_id,
            "current_status": alert.status,
            "current_status_display": AlertStatus.get_display_name(alert.status),
            "processed_by": alert.processed_by,
            "processed_at": alert.processed_at,
            "processing_notes": alert.processing_notes,
            "history": processing_history,
            "total_steps": len(processing_history)
        }
        
        logger.info(f"✅ 成功获取预警 {alert_id} 的处理历史，共 {len(processing_history)} 条记录")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警处理历史失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警处理历史失败: {str(e)}")


@router.get("/by-status/{status}", description="根据处理状态获取预警列表")
def get_alerts_by_processing_status(
    status: int,
    limit: int = Query(default=100, ge=1, le=1000, description="每页数量"),
    offset: int = Query(default=0, ge=0, description="偏移量"),
    db: Session = Depends(get_db)
):
    """
    根据处理状态获取预警列表
    status: 1-待处理, 2-处理中, 3-已处理, 4-已归档, 5-误报
    """
    try:
        # 验证状态值
        if status not in [1, 2, 3, 4, 5]:
            raise HTTPException(status_code=400, detail=f"无效的状态值: {status}")
        
        # 查询预警列表
        query = db.query(Alert).filter(Alert.status == status)
        total = query.count()
        
        alerts = query.order_by(desc(Alert.alert_time)).offset(offset).limit(limit).all()
        
        # 转换为响应格式
        alert_list = []
        for alert in alerts:
            alert_dict = {
                "alert_id": alert.alert_id,
                "alert_name": alert.alert_name,
                "alert_type": alert.alert_type,
                "camera_name": alert.camera_name,
                "location": alert.location,
                "alert_time": alert.alert_time,
                "status": alert.status,
                "status_display": AlertStatus.get_display_name(alert.status),
                "processed_by": alert.processed_by,
                "processed_at": alert.processed_at
            }
            alert_list.append(alert_dict)
        
        result = {
            "alerts": alert_list,
            "total": total,
            "limit": limit,
            "offset": offset,
            "status": status,
            "status_display": AlertStatus.get_display_name(status)
        }
        
        logger.info(f"✅ 成功获取状态为 {AlertStatus.get_display_name(status)} 的预警列表，共 {total} 条")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取预警列表失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警列表失败: {str(e)}")


def _is_valid_status_transition(current_status: int, target_status: int) -> bool:
    """验证状态转换的合法性"""
    # 允许同状态转换（用于更新处理意见、处理人等信息）
    if current_status == target_status:
        return True
    
    # 定义合法的状态转换路径
    valid_transitions = {
        AlertStatus.PENDING: [AlertStatus.PROCESSING, AlertStatus.FALSE_ALARM, AlertStatus.ARCHIVED],
        AlertStatus.PROCESSING: [AlertStatus.RESOLVED, AlertStatus.ARCHIVED, AlertStatus.FALSE_ALARM],
        AlertStatus.RESOLVED: [AlertStatus.ARCHIVED, AlertStatus.PROCESSING],  # 允许重新处理
        AlertStatus.ARCHIVED: [AlertStatus.PROCESSING],  # 允许从归档恢复
        AlertStatus.FALSE_ALARM: [AlertStatus.PROCESSING]  # 允许从误报恢复
    }
    
    allowed_next_states = valid_transitions.get(current_status, [])
    return target_status in allowed_next_states


def _get_action_type_from_status_change(from_status: int, to_status: int) -> int:
    """根据状态变化确定动作类型"""
    from app.models.alert import ProcessingActionType
    
    # 状态转换映射到动作类型
    status_action_map = {
        (AlertStatus.PENDING, AlertStatus.PROCESSING): ProcessingActionType.START_PROCESSING,
        (AlertStatus.PROCESSING, AlertStatus.RESOLVED): ProcessingActionType.FINISH_PROCESSING,
        (AlertStatus.RESOLVED, AlertStatus.ARCHIVED): ProcessingActionType.ARCHIVE,
        (AlertStatus.PENDING, AlertStatus.FALSE_ALARM): ProcessingActionType.MARK_FALSE_ALARM,
        (AlertStatus.PROCESSING, AlertStatus.FALSE_ALARM): ProcessingActionType.MARK_FALSE_ALARM,
        (AlertStatus.ARCHIVED, AlertStatus.PROCESSING): ProcessingActionType.REOPEN,
        (AlertStatus.FALSE_ALARM, AlertStatus.PROCESSING): ProcessingActionType.REOPEN,
    }
    
    return status_action_map.get((from_status, to_status), ProcessingActionType.UPDATE_NOTES)


def _get_action_description(action_type: int, from_status: int, to_status: int) -> str:
    """获取动作描述"""
    from app.models.alert import ProcessingActionType
    
    descriptions = {
        ProcessingActionType.START_PROCESSING: "开始处理预警",
        ProcessingActionType.FINISH_PROCESSING: "完成预警处理",
        ProcessingActionType.ARCHIVE: "归档预警",
        ProcessingActionType.MARK_FALSE_ALARM: "标记为误报",
        ProcessingActionType.REOPEN: "重新处理预警",
        ProcessingActionType.UPDATE_NOTES: "更新处理意见"
    }
    
    return descriptions.get(action_type, f"状态更新: {AlertStatus.get_display_name(from_status)} -> {AlertStatus.get_display_name(to_status)}")


@router.get("/{alert_id}/processing-records", description="获取预警的所有处理记录")
def get_alert_processing_records(
    alert_id: int,
    db: Session = Depends(get_db)
):
    """
    获取预警的所有处理记录（从alert_processing_records表）
    """
    try:
        # 1. 验证预警是否存在
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail=f"预警记录不存在: {alert_id}")
        
        # 2. 查询处理记录
        from app.models.alert import AlertProcessingRecord
        records = db.query(AlertProcessingRecord)\
                    .filter(AlertProcessingRecord.alert_id == alert_id)\
                    .order_by(AlertProcessingRecord.created_at.desc())\
                    .all()
        
        # 3. 转换为响应格式
        processing_records = []
        for record in records:
            processing_records.append({
                "record_id": record.record_id,
                "action_type": record.action_type,
                "action_display": record.action_display,
                "from_status": record.from_status,
                "from_status_display": record.from_status_display,
                "to_status": record.to_status,
                "to_status_display": record.to_status_display,
                "operator_name": record.operator_name,
                "operator_role": record.operator_role,
                "operator_department": record.operator_department,
                "notes": record.notes,
                "processing_duration": record.processing_duration,
                "priority_level": record.priority_level,
                "priority_display": record.priority_display,
                "is_automated": record.is_automated,
                "created_at": record.created_at,
                "updated_at": record.updated_at
            })
        
        result = {
            "alert_id": alert_id,
            "total_records": len(processing_records),
            "processing_records": processing_records
        }
        
        logger.info(f"✅ 成功获取预警 {alert_id} 的处理记录，共 {len(processing_records)} 条")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取处理记录失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取处理记录失败: {str(e)}")


@router.put("/{alert_id}/status", description="更新预警状态并创建处理记录")
def update_alert_status(
    alert_id: int,
    alert_update: AlertUpdate,
    db: Session = Depends(get_db),
    user: Optional[UserInfo] = Depends(get_current_user_optional)
):
    """
    更新预警状态 - 前端确认处理按钮调用的API
    同时更新alerts表和自动创建alert_processing_records记录
    
    操作人信息从JWT Token中自动获取，无需前端传递
    """
    try:
        # 从JWT Token获取当前用户信息
        operator_name = user.userName if user else (alert_update.processed_by or "系统操作")
        operator_dept = user.deptName if user else "未知部门"
        
        logger.info(f"🔄 开始处理预警状态更新: alert_id={alert_id}, status={alert_update.status}, 操作人={operator_name}")
        
        # 1. 查找预警记录
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            logger.error(f"预警记录不存在: {alert_id}")
            raise HTTPException(status_code=404, detail=f"预警记录不存在: {alert_id}")
        
        # 2. 记录原状态
        original_status = alert.status
        logger.info(f"预警 {alert_id} 状态变更: {original_status} -> {alert_update.status}")
        
        # 3. 验证状态转换的合法性
        if not _is_valid_status_transition(original_status, alert_update.status):
            error_msg = f"不允许的状态转换: {AlertStatus.get_display_name(original_status)} -> {AlertStatus.get_display_name(alert_update.status)}"
            logger.error(error_msg)
            raise HTTPException(status_code=400, detail=error_msg)
        
        # 4. 更新预警基本信息（使用当前登录用户信息）
        alert.status = alert_update.status
        alert.processed_by = operator_name  # 使用当前登录用户
        alert.processing_notes = alert_update.processing_notes
        alert.processed_at = datetime.now()
        alert.updated_at = datetime.now()
        
        # 5. 创建处理记录 - 关键步骤！
        from app.models.alert import AlertProcessingRecord, ProcessingActionType
        
        # 根据状态变化确定动作类型
        action_type = _get_action_type_from_status_change(original_status, alert_update.status)
        
        processing_record = AlertProcessingRecord(
            alert_id=alert_id,
            action_type=action_type,
            from_status=original_status,
            to_status=alert_update.status,
            operator_name=operator_name,  # 使用当前登录用户
            operator_role="处理员",
            operator_department=operator_dept,  # 使用当前登录用户的部门
            notes=alert_update.processing_notes,
            priority_level=0,
            is_automated=False,
            created_at=datetime.now()
        )
        
        logger.info(f"📝 创建处理记录: action_type={action_type}, operator={operator_name}, dept={operator_dept}")
        
        # 6. 同时更新JSON格式的process字段（兼容性）
        action_desc = _get_action_description(action_type, original_status, alert_update.status)
        alert.add_process_step(
            step=action_desc,
            desc=alert_update.processing_notes or action_desc,
            operator=operator_name  # 使用当前登录用户
        )
        
        # 7. 保存到数据库
        db.add(processing_record)
        db.commit()
        
        logger.info(f"✅ 成功保存处理记录到数据库: record_id={processing_record.record_id}")
        
        # 8. 刷新获取最新数据
        db.refresh(alert)
        db.refresh(processing_record)
        
        # 9. 返回处理结果（前端期望的格式）
        result = {
            "code": 0,
            "msg": "success", 
            "data": {
                "success": True,
                "message": f"预警 {alert_id} 状态更新成功",
                "alert_id": alert_id,
                "status_change": {
                    "from": original_status,
                    "from_display": AlertStatus.get_display_name(original_status),
                    "to": alert_update.status,
                    "to_display": AlertStatus.get_display_name(alert_update.status)
                },
                "processing_record": {
                    "record_id": processing_record.record_id,
                    "action_type": processing_record.action_type,
                    "action_display": processing_record.action_display,
                    "created_at": processing_record.created_at,
                    "operator": processing_record.operator_name
                },
                "updated_alert": {
                    "alert_id": alert.alert_id,
                    "status": alert.status,
                    "status_display": AlertStatus.get_display_name(alert.status),
                    "processed_by": alert.processed_by,
                    "processed_at": alert.processed_at,
                    "processing_notes": alert.processing_notes
                }
            }
        }
        
        logger.info(f"✅ 预警状态更新完成: {alert_id}")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新预警状态失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"更新预警状态失败: {str(e)}")


@router.put("/batch-update", description="批量更新预警状态并创建处理记录") 
def batch_update_alert_status(
    batch_request: dict,
    db: Session = Depends(get_db),
    user: Optional[UserInfo] = Depends(get_current_user_optional)
):
    """
    批量更新预警状态 - 前端批量处理调用的API
    同时更新alerts表和创建alert_processing_records记录
    
    操作人信息从JWT Token中自动获取，无需前端传递
    """
    try:
        alert_ids = batch_request.get("alert_ids", [])
        if not alert_ids:
            raise HTTPException(status_code=400, detail="缺少预警ID列表")
        
        # 从JWT Token获取当前用户信息
        operator_name = user.userName if user else (batch_request.get("processed_by") or "系统操作")
        operator_dept = user.deptName if user else "未知部门"
        
        logger.info(f"🔄 开始批量处理预警: {len(alert_ids)} 个预警, 操作人={operator_name}")
        
        # 提取更新数据
        status = batch_request.get("status")
        processing_notes = batch_request.get("processing_notes")
        
        if status is None:
            raise HTTPException(status_code=400, detail="缺少状态参数")
        
        success_count = 0
        failure_count = 0
        results = []
        
        for alert_id in alert_ids:
            try:
                # 查找预警记录
                alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
                if not alert:
                    results.append({
                        "alert_id": alert_id,
                        "success": False,
                        "error": f"预警记录不存在: {alert_id}"
                    })
                    failure_count += 1
                    continue
                
                # 记录原状态
                original_status = alert.status
                
                # 验证状态转换
                if not _is_valid_status_transition(original_status, status):
                    results.append({
                        "alert_id": alert_id,
                        "success": False,
                        "error": f"不允许的状态转换: {AlertStatus.get_display_name(original_status)} -> {AlertStatus.get_display_name(status)}"
                    })
                    failure_count += 1
                    continue
                
                # 更新预警（使用当前登录用户信息）
                alert.status = status
                alert.processed_by = operator_name  # 使用当前登录用户
                alert.processing_notes = processing_notes
                alert.processed_at = datetime.now()
                alert.updated_at = datetime.now()
                
                # 创建处理记录
                from app.models.alert import AlertProcessingRecord, ProcessingActionType
                action_type = _get_action_type_from_status_change(original_status, status)
                
                processing_record = AlertProcessingRecord(
                    alert_id=alert_id,
                    action_type=action_type,
                    from_status=original_status,
                    to_status=status,
                    operator_name=operator_name,  # 使用当前登录用户
                    operator_role="处理员",
                    operator_department=operator_dept,  # 使用当前登录用户的部门
                    notes=processing_notes,
                    priority_level=0,
                    is_automated=False,
                    created_at=datetime.now()
                )
                
                # 更新JSON字段
                action_desc = _get_action_description(action_type, original_status, status)
                alert.add_process_step(
                    step=action_desc,
                    desc=processing_notes or action_desc,
                    operator=operator_name  # 使用当前登录用户
                )
                
                db.add(processing_record)
                
                results.append({
                    "alert_id": alert_id,
                    "success": True,
                    "processing_record_id": processing_record.record_id,
                    "status_change": f"{AlertStatus.get_display_name(original_status)} -> {AlertStatus.get_display_name(status)}"
                })
                success_count += 1
                
                logger.info(f"✅ 批量处理成功: alert_id={alert_id}")
                
            except Exception as e:
                logger.error(f"批量处理单个预警失败: alert_id={alert_id}, error={str(e)}")
                results.append({
                    "alert_id": alert_id,
                    "success": False,
                    "error": str(e)
                })
                failure_count += 1
        
        # 提交所有更改
        db.commit()
        
        # 返回结果
        result = {
            "code": 0 if failure_count == 0 else -1,
            "msg": "success" if failure_count == 0 else f"部分失败: {failure_count}个失败",
            "data": {
                "total": len(alert_ids),
                "success_count": success_count,
                "failure_count": failure_count,
                "results": results
            }
        }
        
        logger.info(f"✅ 批量更新预警状态完成: {success_count}成功, {failure_count}失败")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"批量更新预警状态失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"批量更新预警状态失败: {str(e)}")


class BatchDeleteAlertsRequest(BaseModel):
    """批量删除预警请求模型"""
    alert_ids: List[int]


@router.post("/batch-delete", summary="批量删除预警")
async def batch_delete_alerts(
    request: BatchDeleteAlertsRequest,
    db: Session = Depends(get_db)
):
    """
    批量删除预警记录
    
    Args:
        request: 包含预警ID列表的请求体
        db: 数据库会话
        
    Returns:
        批量删除结果
    """
    try:
        alert_ids = request.alert_ids
        if not alert_ids:
            raise HTTPException(status_code=400, detail="预警ID列表不能为空")
            
        logger.info(f"开始批量删除预警: {alert_ids}")
        
        # 查询要删除的预警记录
        alerts_to_delete = db.query(Alert).filter(Alert.alert_id.in_(alert_ids)).all()
        found_alert_ids = [alert.alert_id for alert in alerts_to_delete]
        not_found_ids = [alert_id for alert_id in alert_ids if alert_id not in found_alert_ids]
        
        deleted_count = 0
        if alerts_to_delete:
            # 删除预警记录（会级联删除相关的处理记录）
            for alert in alerts_to_delete:
                db.delete(alert)
                deleted_count += 1
                logger.debug(f"删除预警记录: alert_id={alert.alert_id}")
            
            db.commit()
            logger.info(f"批量删除预警完成，共删除 {deleted_count} 条记录")
        
        # 构建响应信息
        if deleted_count == 0:
            if not_found_ids:
                message = f"所选预警记录不存在，无法删除。未找到的ID: {not_found_ids}"
            else:
                message = "没有找到可删除的预警记录"
        elif not_found_ids:
            message = f"成功删除 {deleted_count} 条记录，{len(not_found_ids)} 条记录未找到（ID: {not_found_ids}）"
        else:
            message = f"成功删除 {deleted_count} 条预警记录"
            
        return {
            "code": 0,
            "msg": message,
            "data": {
                "deleted_count": deleted_count,
                "not_found_ids": not_found_ids,
                "total_requested": len(alert_ids)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"批量删除预警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"批量删除预警失败: {str(e)}")


@router.delete("/{alert_id}", summary="删除单个预警")
async def delete_alert(
    alert_id: int,
    db: Session = Depends(get_db)
):
    """
    删除单个预警记录
    
    Args:
        alert_id: 预警ID
        db: 数据库会话
        
    Returns:
        删除结果
    """
    try:
        logger.info(f"开始删除预警: {alert_id}")
        
        # 查询要删除的预警记录
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail=ALERT_NOT_FOUND_MSG)
        
        # 删除预警记录（会级联删除相关的处理记录）
        db.delete(alert)
        db.commit()
        
        logger.info(f"删除预警成功: alert_id={alert_id}")
        
        return {
            "code": 0,
            "msg": "预警删除成功",
            "data": {
                "alert_id": alert_id
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除预警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除预警失败: {str(e)}")


@router.post("/{alert_id}/false-alarm", summary="标记预警为误报")
async def mark_alert_as_false_alarm(
    alert_id: int,
    review_notes: str = Query(..., description="复判意见"),
    reviewer_name: str = Query(..., description="复判人员姓名"),
    db: Session = Depends(get_db)
):
    """
    标记预警为误报，并创建复判记录
    
    Args:
        alert_id: 预警ID
        review_notes: 复判意见
        reviewer_name: 复判人员姓名
        db: 数据库会话
        
    Returns:
        误报处理结果
    """
    try:
        logger.info(f"开始标记预警为误报: alert_id={alert_id}, reviewer={reviewer_name}")
        
        # 查询预警记录
        alert = db.query(Alert).filter(Alert.alert_id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail=ALERT_NOT_FOUND_MSG)
        
        # 检查预警是否已经是误报状态
        if alert.status == AlertStatus.FALSE_ALARM:
            return {
                "code": 0,
                "msg": "预警已经是误报状态",
                "data": {
                    "alert_id": alert_id,
                    "status": alert.status,
                    "status_display": alert.status_display
                }
            }
        
        # 检查预警状态：只有待处理状态才能标记为误报
        if alert.status != AlertStatus.PENDING:
            status_names = {
                AlertStatus.PROCESSING: "处理中",
                AlertStatus.RESOLVED: "已处理",
                AlertStatus.ARCHIVED: "已归档"
            }
            current_status_name = status_names.get(alert.status, alert.status_display)
            raise HTTPException(
                status_code=400,
                detail=f"只有待处理状态的预警才能标记为误报，当前状态为：{current_status_name}"
            )
        
        # 更新预警状态为误报
        old_status = alert.status
        alert.status = AlertStatus.FALSE_ALARM
        alert.processed_at = datetime.utcnow()
        alert.processed_by = reviewer_name
        alert.processing_notes = f"标记为误报：{review_notes}"
        
        # 添加处理流程步骤
        alert.add_process_step("标记误报", f"复判人员 {reviewer_name} 标记为误报：{review_notes}", reviewer_name)
        
        # 创建复判记录
        from app.db.review_record_dao import ReviewRecordDAO
        review_dao = ReviewRecordDAO(db)
        review_record = review_dao.create_review_record(
            alert_id=alert_id,
            review_type="manual",
            reviewer_name=reviewer_name,
            review_notes=review_notes
        )
        
        if not review_record:
            logger.warning(f"创建复判记录失败: alert_id={alert_id}")
        
        # 创建处理记录
        from app.models.alert import AlertProcessingRecord, ProcessingActionType
        processing_record = AlertProcessingRecord(
            alert_id=alert_id,
            action_type=ProcessingActionType.MARK_FALSE_ALARM,
            from_status=old_status,
            to_status=AlertStatus.FALSE_ALARM,
            operator_name=reviewer_name,
            operator_role="复判人员",
            notes=review_notes,
            created_at=datetime.utcnow()
        )
        db.add(processing_record)
        
        db.commit()
        
        logger.info(f"标记误报成功: alert_id={alert_id}, reviewer={reviewer_name}")
        
        return {
            "code": 0,
            "msg": "预警已标记为误报",
            "data": {
                "alert_id": alert_id,
                "status": alert.status,
                "status_display": alert.status_display,
                "review_record_id": review_record.review_id if review_record else None,
                "processed_at": alert.processed_at,
                "processed_by": alert.processed_by
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"标记误报失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"标记误报失败: {str(e)}")


@router.post("/batch-false-alarm", summary="批量标记预警为误报")
async def batch_mark_alerts_as_false_alarm(
    request: BatchDeleteAlertsRequest,  # 复用批量删除的请求模型
    review_notes: str = Query(..., description="复判意见"),
    reviewer_name: str = Query(..., description="复判人员姓名"),
    db: Session = Depends(get_db)
):
    """
    批量标记预警为误报
    
    Args:
        request: 包含预警ID列表的请求体
        review_notes: 复判意见
        reviewer_name: 复判人员姓名
        db: 数据库会话
        
    Returns:
        批量误报处理结果
    """
    try:
        alert_ids = request.alert_ids
        if not alert_ids:
            raise HTTPException(status_code=400, detail="预警ID列表不能为空")
            
        logger.info(f"开始批量标记误报: {alert_ids}, reviewer={reviewer_name}")
        
        # 查询要处理的预警记录
        alerts_to_process = db.query(Alert).filter(Alert.alert_id.in_(alert_ids)).all()
        found_alert_ids = [alert.alert_id for alert in alerts_to_process]
        not_found_ids = [alert_id for alert_id in alert_ids if alert_id not in found_alert_ids]
        
        processed_count = 0
        already_false_alarm_count = 0
        skipped_non_pending_count = 0
        
        for alert in alerts_to_process:
            if alert.status == AlertStatus.FALSE_ALARM:
                already_false_alarm_count += 1
                continue
            
            # 检查预警状态：只有待处理状态才能标记为误报
            if alert.status != AlertStatus.PENDING:
                skipped_non_pending_count += 1
                logger.warning(f"跳过非待处理状态的预警: alert_id={alert.alert_id}, status={alert.status}")
                continue
                
            # 更新预警状态为误报
            old_status = alert.status
            alert.status = AlertStatus.FALSE_ALARM
            alert.processed_at = datetime.utcnow()
            alert.processed_by = reviewer_name
            alert.processing_notes = f"批量标记为误报：{review_notes}"
            
            # 添加处理流程步骤
            alert.add_process_step("批量标记误报", f"复判人员 {reviewer_name} 批量标记为误报：{review_notes}", reviewer_name)
            
            # 创建复判记录
            from app.db.review_record_dao import ReviewRecordDAO
            review_dao = ReviewRecordDAO(db)
            review_record = review_dao.create_review_record(
                alert_id=alert.alert_id,
                review_type="manual",
                reviewer_name=reviewer_name,
                review_notes=review_notes
            )
            
            # 创建处理记录
            from app.models.alert import AlertProcessingRecord, ProcessingActionType
            processing_record = AlertProcessingRecord(
                alert_id=alert.alert_id,
                action_type=ProcessingActionType.MARK_FALSE_ALARM,
                from_status=old_status,
                to_status=AlertStatus.FALSE_ALARM,
                operator_name=reviewer_name,
                operator_role="复判人员",
                notes=review_notes,
                created_at=datetime.utcnow()
            )
            db.add(processing_record)
            
            processed_count += 1
        
        db.commit()
        
        logger.info(f"批量标记误报完成，共处理 {processed_count} 条记录")
        
        # 构建响应信息
        message_parts = []
        if processed_count > 0:
            message_parts.append(f"成功标记 {processed_count} 条预警为误报")
        if already_false_alarm_count > 0:
            message_parts.append(f"{already_false_alarm_count} 条预警已经是误报状态")
        if skipped_non_pending_count > 0:
            message_parts.append(f"{skipped_non_pending_count} 条预警因非待处理状态被跳过")
        if not_found_ids:
            message_parts.append(f"{len(not_found_ids)} 条预警记录未找到")
            
        message = "；".join(message_parts) if message_parts else "没有找到可处理的预警记录"
        
        return {
            "code": 0,
            "msg": message,
            "data": {
                "processed_count": processed_count,
                "already_false_alarm_count": already_false_alarm_count,
                "skipped_non_pending_count": skipped_non_pending_count,
                "not_found_ids": not_found_ids,
                "total_requested": len(alert_ids)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"批量标记误报失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"批量标记误报失败: {str(e)}")


# ========== JWT用户信息使用示例接口 ==========
# 以下接口演示如何在实际业务中使用JWT Token解析功能

@router.get("/user/my-alerts", response_model=Dict[str, Any], description="获取当前用户的预警列表（示例）")
async def get_my_alerts(
    user: UserInfo = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1, description="页码"),
    limit: int = Query(10, ge=1, le=100, description="每页数量")
):
    """
    【示例接口1：强制要求用户登录】
    获取当前登录用户创建的预警列表
    
    使用场景：
    - 个人中心查看我的预警
    - 需要用户认证的接口
    
    技术要点：
    - 使用 Depends(get_current_user) 强制要求用户登录
    - Token无效时自动返回401错误
    - 可以直接使用 user.userId, user.userName 等属性
    """
    try:
        logger.info(f"用户 {user.userName}(ID:{user.userId}) 查询自己的预警列表")
        
        # 这里可以根据用户ID查询预警
        # 示例：查询该用户所在租户的预警
        query = db.query(Alert)
        
        # 如果有租户ID，按租户过滤（多租户场景）
        if user.tenantId:
            logger.debug(f"按租户ID过滤: {user.tenantId}")
            # query = query.filter(Alert.tenant_id == user.tenantId)
        
        # 分页
        total = query.count()
        alerts = query.order_by(desc(Alert.alert_time)).offset((page - 1) * limit).limit(limit).all()
        
        return {
            "code": 0,
            "msg": "success",
            "data": {
                "user_info": {
                    "userId": user.userId,
                    "userName": user.userName,
                    "deptName": user.deptName,
                    "tenantId": user.tenantId
                },
                "alerts": [
                    {
                        "id": alert.id,
                        "alert_name": alert.alert_name,
                        "alert_time": alert.alert_time.isoformat() if alert.alert_time else None,
                        "alert_level": alert.alert_level,
                        "status": alert.status
                    }
                    for alert in alerts
                ],
                "pagination": {
                    "page": page,
                    "limit": limit,
                    "total": total,
                    "pages": math.ceil(total / limit) if limit > 0 else 0
                }
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取用户预警列表失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警列表失败: {str(e)}")


@router.get("/user/info", response_model=Dict[str, Any], description="获取当前用户信息（示例）")
async def get_current_user_info(
    user: UserInfo = Depends(get_current_user)
):
    """
    【示例接口2：获取当前用户信息】
    返回JWT Token中的所有用户信息
    
    使用场景：
    - 前端获取当前登录用户信息
    - 用户个人信息展示
    
    技术要点：
    - 直接返回UserInfo对象中的所有字段
    - 演示如何访问Token中的所有字段
    """
    logger.info(f"用户 {user.userName}(ID:{user.userId}) 查询个人信息")
    
    return {
        "code": 0,
        "msg": "success",
        "data": {
            "loginType": user.loginType,
            "loginId": user.loginId,
            "clientid": user.clientid,
            "tenantId": user.tenantId,
            "userId": user.userId,
            "userName": user.userName,
            "deptId": user.deptId,
            "deptName": user.deptName,
            "deptCategory": user.deptCategory
        }
    }


@router.get("/public/alerts", response_model=Dict[str, Any], description="公开预警列表（可选登录，示例）")
async def get_public_alerts(
    user: Optional[UserInfo] = Depends(get_current_user_optional),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1, description="页码"),
    limit: int = Query(10, ge=1, le=100, description="每页数量")
):
    """
    【示例接口3：可选用户登录】
    获取预警列表，支持匿名访问，但登录用户可以看到更多信息
    
    使用场景：
    - 公开API，但登录用户有额外功能
    - 个性化推荐（登录用户基于偏好）
    
    技术要点：
    - 使用 Depends(get_current_user_optional) 允许匿名访问
    - user 可能为 None，需要判断
    - 根据是否登录提供不同的数据
    """
    if user:
        logger.info(f"用户 {user.userName}(ID:{user.userId}) 访问公开预警列表（已登录）")
    else:
        logger.info("匿名用户访问公开预警列表")
    
    try:
        query = db.query(Alert)
        
        # 如果用户已登录且有租户ID，可以优先显示本租户的预警
        if user and user.tenantId:
            logger.debug(f"已登录用户，租户ID: {user.tenantId}")
            # 可以添加租户相关的过滤或排序逻辑
        
        # 分页
        total = query.count()
        alerts = query.order_by(desc(Alert.alert_time)).offset((page - 1) * limit).limit(limit).all()
        
        response_data = {
            "alerts": [
                {
                    "id": alert.id,
                    "alert_name": alert.alert_name,
                    "alert_time": alert.alert_time.isoformat() if alert.alert_time else None,
                    "alert_level": alert.alert_level,
                    "status": alert.status
                }
                for alert in alerts
            ],
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total,
                "pages": math.ceil(total / limit) if limit > 0 else 0
            }
        }
        
        # 如果用户已登录，添加用户信息
        if user:
            response_data["user_info"] = {
                "userId": user.userId,
                "userName": user.userName,
                "authenticated": True
            }
        else:
            response_data["user_info"] = {
                "authenticated": False
            }
        
        return {
            "code": 0,
            "msg": "success",
            "data": response_data
        }
        
    except Exception as e:
        logger.error(f"获取公开预警列表失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取预警列表失败: {str(e)}")


@router.post("/alerts/{alert_id}/claim", response_model=Dict[str, Any], description="认领预警（示例）")
async def claim_alert(
    alert_id: int,
    user: UserInfo = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    【示例接口4：使用用户信息更新数据】
    认领预警，将当前用户设置为预警的处理人
    
    使用场景：
    - 预警处理流程
    - 需要记录操作人信息
    
    技术要点：
    - 使用用户ID和用户名更新数据库
    - 演示如何在业务逻辑中使用用户信息
    """
    try:
        logger.info(f"用户 {user.userName}(ID:{user.userId}) 尝试认领预警 {alert_id}")
        
        # 查询预警
        alert = db.query(Alert).filter(Alert.id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail=ALERT_NOT_FOUND_MSG)
        
        # 检查预警状态
        if alert.status != AlertStatus.PENDING.value:
            raise HTTPException(status_code=400, detail="只能认领待处理状态的预警")
        
        # 更新预警状态和处理人信息
        alert.status = AlertStatus.PROCESSING.value
        # 假设Alert模型有这些字段（实际使用时需要确认）
        # alert.handler_id = user.userId
        # alert.handler_name = user.userName
        # alert.claim_time = datetime.now()
        
        db.commit()
        
        logger.info(f"预警 {alert_id} 已被用户 {user.userName} 认领")
        
        return {
            "code": 0,
            "msg": "预警认领成功",
            "data": {
                "alert_id": alert_id,
                "handler": {
                    "userId": user.userId,
                    "userName": user.userName,
                    "deptName": user.deptName
                },
                "status": alert.status
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"认领预警失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"认领预警失败: {str(e)}")


@router.get("/user/context-demo", response_model=Dict[str, Any], description="使用工具函数获取用户信息（示例）")
async def get_user_context_demo(request: Request):
    """
    【示例接口5：使用工具函数】
    演示如何使用 auth_helper 中的工具函数快速获取用户信息
    
    使用场景：
    - 不想使用依赖注入
    - 在中间件或其他地方需要获取用户信息
    
    技术要点：
    - 使用 get_user_context() 获取完整用户上下文
    - 使用 get_user_id_from_request() 等快捷函数
    """
    # 方式1：获取完整用户上下文
    user_context = get_user_context(request)
    
    # 方式2：单独获取某个字段
    user_id = get_user_id_from_request(request)
    user_name = get_user_name_from_request(request)
    
    logger.info(f"工具函数获取用户信息: userId={user_id}, userName={user_name}")
    
    return {
        "code": 0,
        "msg": "success",
        "data": {
            "method": "auth_helper工具函数",
            "user_context": user_context,
            "individual_fields": {
                "userId": user_id,
                "userName": user_name
            }
        }
    }


@router.get("/forward-statistics", summary="获取报警转发统计")
async def get_alert_forward_statistics(
    time_range: str = Query("7d", description="时间范围: 7d/30d"),
    db: Session = Depends(get_db)
):
    """
    获取报警转发统计（真实数据版本）

    统计通过 RabbitMQ 发布和 SSE 通知的报警数量

    返回格式:
    {
        "code": 0,
        "msg": "success",
        "data": {
            "forward_counts": [10, 20, 15, ...],  # 每日转发数量
            "date_labels": ["2024-01-01", "2024-01-02", ...],  # 日期标签
            "total_forwards": 1000,  # 总转发数
            "publish_count": 800,     # RabbitMQ发布数
            "notification_count": 200  # SSE通知数
        }
    }
    """
    try:
        from datetime import timedelta
        from app.db.compensation_dao import AlertPublishLogDAO, AlertNotificationLogDAO

        logger.info(f"收到获取报警转发统计请求: time_range={time_range}")

        # 解析时间范围
        end_date = datetime.now()
        if time_range == "30d":
            days = 30
        elif time_range == "7d":
            days = 7
        else:
            days = 7

        start_date = end_date - timedelta(days=days)

        # 获取发布统计（RabbitMQ）
        publish_stats = AlertPublishLogDAO.get_daily_statistics(db, start_date, end_date)
        publish_total = AlertPublishLogDAO.get_total_count(db, start_date, end_date)

        # 获取通知统计（SSE）
        notification_stats = AlertNotificationLogDAO.get_daily_statistics(db, start_date, end_date)

        # 构建日期标签和每日统计数据
        date_labels = []
        forward_counts = []

        # 创建日期映射字典
        publish_dict = {item["date"]: item["count"] for item in publish_stats}
        notification_dict = {item["date"]: item["count"] for item in notification_stats}

        # 获取所有有数据的日期并排序
        all_dates = set(publish_dict.keys()) | set(notification_dict.keys())
        sorted_dates = sorted(all_dates)

        for date_str in sorted_dates:
            date_labels.append(date_str)
            # 每日转发数 = 发布数 + 通知数
            daily_count = publish_dict.get(date_str, 0) + notification_dict.get(date_str, 0)
            forward_counts.append(daily_count)

        # 如果没有数据，生成空的时间序列
        if not date_labels:
            for i in range(days):
                date = start_date + timedelta(days=i)
                date_labels.append(date.strftime("%Y-%m-%d"))
                forward_counts.append(0)

        total_forwards = publish_total

        return {
            "code": 0,
            "msg": "success",
            "data": {
                "forward_counts": forward_counts,
                "date_labels": date_labels,
                "total_forwards": total_forwards,
                "publish_count": publish_total,
                "notification_count": sum(notification_dict.values())
            }
        }

    except Exception as e:
        logger.error(f"获取报警转发统计失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取报警转发统计失败: {str(e)}")
